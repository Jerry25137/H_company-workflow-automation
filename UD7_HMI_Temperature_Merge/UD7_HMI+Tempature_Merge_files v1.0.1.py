# -*- coding: utf-8 -*-
"""
Created on Tue Jan 21 14:25:46 2025

@author: USER
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.drawing.line import LineProperties
import win32com.client as win32

from tkinter import filedialog, messagebox

data_units = {"FREQ":"Frequency [Hz]", "IFB":"Current [mA]", "VFB":"Power [%]", "TEMP":"Tempature [℃] (x10)"}

linetypes = ["solid","sysDash", "sysDashDot", "sysDashDotDot", "sysDot",
             "dash", "dashDot", "dot", "lgDash", "lgDashDot", "lgDashDotDot",]

# 54種循環顏色
colors = ["4F81BD", "C0504D", "9BBB59", "8064A2", "4BACC6", "F79646", "2C4D75", "772C2A", "5F7530", "4D3B62", 
          "276A7C", "B65708", "729ACA", "CD7371", "AFC97A", "9983B5", "6FBDD1", "F9AB6B", "3A679C", "9F3B38", 
          "7E9D40", "664F83", "358EA6", "F3740B", "95B3D7", "D99694", "C3D69B", "B3A2C7", "93CDDD", "FAC090", 
          "254061", "632523", "4F6228", "403152", "215968", "984807", "84A7D1", "D38482", "B9CF8B", "A692BE", 
          "81C5D7", "F9B67E", "335A88", "8B3431", "6F8938", "594573", "2E7C91", "D56509", "A7C0DE", "DFA8A6", 
          "CDDDAC", "BFB2D0", "A5D6E2", "FBCBA3", ]

# 讀取XLSX檔案
def read_xlsx(f_path):
    workbook = openpyxl.load_workbook(f_path)
    sheet = workbook.active
    
    rows = []
    for row in sheet.iter_rows():
        # 取前 5 欄，並將 None 替換為空字串
        values = [cell.value if cell.value is not None else "" for cell in row[:5]]
        if any(values):  # 跳過完全空白的行
            rows.append(values)
    
    return rows

def Merge_data(A, B):
    # 合併標題時間列
    headers = ["time"] + A[0][1:] + B[0][1:]
    body = {}

    # 初始化資料結構，使用完整標題長度確保對齊
    default_row = ["" for _ in range(len(headers))]

    # 輸入 A 資料
    for row in A[1:]:
        freq = row[0]  # freq 已是 datetime.datetime 格式
        full_row = default_row[:]
        full_row[:len(row)] = row
        full_row[0] = freq  # 保留時間格式
        body[freq] = full_row

    # 輸入 B 資料
    for row in B[1:]:
        freq = row[0]  # freq 已是 datetime.datetime 格式
        if freq in body:
            body[freq][len(A[0]):] = row[1:]
        else:
            full_row = default_row[:]
            full_row[0] = freq
            full_row[len(A[0]):] = row[1:]
            body[freq] = full_row

    # 重新排序並且塑造結果
    Merge_data = [headers] + [body[freq] for freq in sorted(body)]

    return Merge_data


# 設定圖表標題格式
def set_chart_title_size(chart, size = 1400):
    paraprops = ParagraphProperties()
    paraprops.defRPr = CharacterProperties(sz=size)

    for para in chart.title.tx.rich.paragraphs:
        para.pPr = paraprops
        
# 建立Excel圖表儲存位置
def Drawing_adress(n):
    adress = []
    n += 1
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        adress.append(chr(remainder + ord('A')))
    return ''.join(reversed(adress))

# Excel 圖表繪製 (資料/線顏色/線型/工作分頁)
def Drawing(A, B, colors, ws):
    # 資料換位置，追頻資料優先
    if "FREQ" in B[0] or "IFB" in B[0] or "VFB" in B[0]:
        C = A
        A = B
        B = C
    
    # 溫度倍率 x10
    for i1 in range(1, len(B)):
        for j1 in range(1, len(B[0])):
            B[i1][j1] = B[i1][j1] * 10
    
    DATA = Merge_data(A, B)

    # XY散佈圖
    chart = LineChart()
    chart.title = "UD7 HMI + Tempature"
    set_chart_title_size(chart, size = 1400)
    chart.style = 13

    # 右Y軸
    chart2 = LineChart()
    chart2.y_axis.axId = 200
    chart2.y_axis.crosses = 'max'
    chart2.y_axis.majorGridlines = None  # 取消格線
    chart2.y_axis.majorTickMark  = 'out' # 刻度在外

    # 標準範圍：全左側Y軸
    R = [[2, len(DATA) + 1]]

    # 分割範圍：增加右側Y軸
    if "FREQ" in A[0]:
        R = [[2, 3], [3, len(DATA[0]) + 1]]
        
        # 左Y軸
        chart.y_axis.title = data_units[DATA[0][1]]
        chart.y_axis.majorGridlines = openpyxl.chart.axis.ChartLines() # 打開格線
        
        # 右Y軸
        if "IFB" in A[0]:
            y2_axis_title = data_units[DATA[0][2]]
            
        if "VFB" in A[0]:
            y2_axis_title += "\n" + data_units[DATA[0][3]]
            
        y2_axis_title += "\n" + data_units["TEMP"]
        
        chart2.y_axis.title = y2_axis_title
    
    elif "IFB" in A[0] or "VFB" in A[0]:
        R = [[2, len(A[0]) + 1], [len(A[0]) + 1, len(DATA[0]) + 1]]
        
        if "IFB" in A[0]:
            y_axis_title = data_units[DATA[0][2]]
            
        if "VFB" in A[0]:
            y_axis_title += "\n" + data_units[DATA[0][3]]
        
        chart.y_axis.title  = y_axis_title
        chart2.y_axis.title = data_units["TEMP"]

    # 建立數據對應顏色標籤
    data_colors = []
    for i1 in range(1, len(DATA[0])):
        if DATA[0][i1] == "FREQ":
            data_colors.append(colors[0])

        elif DATA[0][i1] == "IFB":
            data_colors.append(colors[1])

        elif DATA[0][i1] == "VFB":
            data_colors.append(colors[2])
        
        else:
            data_colors.append(colors[i1 - 1])

    # X軸
    chart.x_axis.title = "Time"
    chart.x_axis.number_format = "h:mm:ss.000"
    x_values = Reference(ws, min_col = 1, max_col = 1, min_row = 2, max_row = len(DATA))

    # 左右Y軸資料合併
    for i2 in range(len(R)):
        for y in range(R[i2][0], R[i2][1]):
            y_values = Reference(ws, min_col = y, min_row = 1, max_row = len(DATA))
            series = Series(y_values, title_from_data = True)
            line_properties = LineProperties(w = 12700, solidFill = data_colors[y - 2])
            series.graphicalProperties.line = line_properties

            if i2 == 0:
                chart.append(series)

            elif i2 == 1:
                chart2.append(series)

    # 設定X軸標籤
    chart.set_categories(x_values)
       
    if len(DATA[0]) >= 3:
        chart += chart2

    # 圖表儲存位置
    adress = Drawing_adress(len(DATA[0])) + str("1") 

    chart.height = 15 # 設置高度
    chart.width  = 17 # 設置寬度

    return DATA, chart, adress

# 使用 pywin32 設定「連接資料點的線」
def set_excel_chart_options(file_path):
    excel = win32.DispatchEx("Excel.Application")  # 使用 DispatchEx 確保背景執行
    excel.Visible = False  # 設定為不可見
    excel.DisplayAlerts = False  # 關閉提示

    try:
        # 開啟文件
        workbook = excel.Workbooks.Open(file_path)
        sheet = workbook.Sheets(1)

        # 取得圖表
        chart_object = sheet.ChartObjects(1)  # 取得第一個圖表
        chart = chart_object.Chart

        # 設定「連接資料點的線」
        chart.DisplayBlanksAs = 3  # 3 表示連接空白資料點 (xlInterpolated)

        # 儲存
        workbook.Save()

    finally:
        # 確保清理資源並關閉 Excel
        workbook.Close(SaveChanges = True)
        excel.Quit()

# 儲存至Excel
def Excel_file(f_path, A, B):
    # 創建一個新的 Excel 工作簿
    wb = Workbook()
    
    # 將資料存進Temp分頁
    ws = wb.active
    ws.title = "HMI+TEMP_merge"
    
    # 繪製Excel圖表
    DATA, chart, adress = Drawing(A, B, colors, ws) 

    # 儲存Excel圖表       
    ws.add_chart(chart, adress)
    
    # 存入資料進Excel
    for row in DATA:
        ws.append(row)
    
    try:
        # 儲存檔案
        save_path = f'{f_path}/UD7_HMI+Tempature_Output.xlsx'
        wb.save(save_path)
        set_excel_chart_options(save_path)

        # 提示保存成功
        messagebox.showinfo("成功", "成功合併檔案：UD7_HMI+Tempature_Output.xlsx")

    except Exception as e:
        print("⚠️檢查 1：讀確認Tempature_Output.xlsx，檔案是否有開啟！")
        messagebox.showerror("錯誤", f"合併檔案時出錯：{str(e)}")

print("選取1：UD7_HMI資料 or 溫度資料")
A_path = filedialog.askopenfilename()
if A_path.endswith('.xlsx'):
    A = read_xlsx(A_path)

print("選取2：UD7_HMI資料 or 溫度資料")    
B_path = filedialog.askopenfilename()
if B_path.endswith('.xlsx'):
    B = read_xlsx(B_path)

print("選取：資料儲存位置！")   
f_path = filedialog.askdirectory()
Excel_file(f_path, A, B)

